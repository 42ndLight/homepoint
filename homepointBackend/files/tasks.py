"""
Celery tasks for async file processing.
Includes the refactored xlsx conversion logic from scripts/xlsx_to_json.py
"""

import json
import sys
import time
import io
from pathlib import Path

import openpyxl
from celery import shared_task
from django.core.files.storage import default_storage
from django.core.cache import cache
from django.db import transaction

from products.models import Category, Inventory, Product, Variant
from .models import ImportHistory

# ── per-sheet column → field key maps ─────────────────────────────────────────
# Keys are lowercased header labels from the spreadsheet.
SHEET_MAPS = {
    "Categories": {
        "id": "id",
        "name": "name",
        "slug": "slug",
        "description": "description",
        "parent id": "parent_id",
    },
    "Products": {
        "product id": "id",
        "name": "name",
        "slug": "slug",
        "description": "description",
        "category id": "category_id",
        "base price (kes)": "base_price",
        "base price": "base_price",
        "is active": "is_active",
    },
    "Variants": {
        "variant id": "id",
        "product id": "product_id",
        "sku": "sku",
        "price (kes)": "price",
        "price": "price",
        "unit type": "unit_type",
        "tax type": "tax_type",
        "attributes": "attributes",
        "item code": "item_code",
        "low stock alert": "stock_threshold",
    },
    "Inventory": {
        "variant id": "variant_id",
        "qty": "quantity",
        "location": "location",
        "last updated": "last_updated",
        "low stock?": "is_low_stock",
    },
}


# Positional fallbacks in case column headers are renamed, missing, or corrupted.
POSITIONAL_MAPS = {
    "Categories": ["id", "name", "slug", "description", "parent_id"],
    "Products": [
        "id",
        "name",
        "category_id",
        "base_price",
        "slug",
        "description",
        "is_active",
    ],
    "Variants": [
        "id",
        "product_id",
        "sku",
        "price",
        "quantity",
        "unit_type",
        "tax_type",
        "attributes",
        "item_code",
        "stock_threshold",
    ],
    "Inventory": ["variant_id", "quantity", "location", "last_updated"],
}


def read_sheet(ws):
    """
    Return list of dicts for a sheet, mapping display headers → field keys
    using SHEET_MAPS with positional fallback mapping. Skips row 2 (the notes/hint row).
    """
    col_map = SHEET_MAPS.get(ws.title, {})
    pos_map = POSITIONAL_MAPS.get(ws.title, [])
    valid_fields = set(col_map.values())

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    headers = []
    for i, h in enumerate(raw_headers):
        h_clean = h.lower()
        if h_clean in col_map:
            headers.append(col_map[h_clean])
        elif h_clean in valid_fields:
            headers.append(h_clean)
        elif i < len(pos_map):
            headers.append(pos_map[i])
        else:
            headers.append(h_clean)

    result = []
    for row in rows[2:]:  # row 0=headers, row 1=notes, data starts at row 2
        # Check if the row has any actual content (ignoring None, empty strings, or just spaces)
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        result.append({headers[i]: cell for i, cell in enumerate(row) if headers[i]})
    return result


# ── type helpers ──────────────────────────────────────────────────────────────


def to_int(val, field=""):
    if val is None or str(val).strip() == "":
        return None
    try:
        return int(val)
    except (TypeError, ValueError):
        raise ValueError(f"Expected integer for '{field}', got: {val!r}")


def to_float(val, field=""):
    if val is None or str(val).strip() == "":
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        raise ValueError(f"Expected number for '{field}', got: {val!r}")


def parse_attributes(raw):
    """'color:grey, size:A4'  →  {'color': 'grey', 'size': 'A4'}"""
    if not raw:
        return {}
    result = {}
    for pair in str(raw).split(","):
        pair = pair.strip()
        if ":" not in pair:
            continue
        key, _, val = pair.partition(":")
        result[key.strip()] = val.strip()
    return result


def slugify(name):
    import re

    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


def parse_bool(val):
    if isinstance(val, bool):
        return val
    return str(val or "true").strip().lower() not in ("false", "0", "no")


# ── main conversion ───────────────────────────────────────────────────────────


def convert(storage_key: str) -> dict:
    """
    Converts XLSX file at xlsx_path to manifest dict matching the expected seed format.
    Handles all sheets: Categories, Products, Variants, Inventory.
    Reads file directly from Django default_storage.
    """
    with default_storage.open(storage_key, 'rb') as f:
        file_bytes = f.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        try:
            missing = {"Categories", "Products", "Variants", "Inventory"} - set(
                wb.sheetnames
            )
            if missing:
                raise ValueError(f"Missing sheets: {', '.join(missing)}")
    
            raw_cats = read_sheet(wb["Categories"])
            raw_prods = read_sheet(wb["Products"])
            raw_vars = read_sheet(wb["Variants"])
            raw_inv = read_sheet(wb["Inventory"])
        finally:
            wb.close()

    # ── categories ────────────────────────────────────────────────────────────
    categories = []
    cat_lookup = {}

    for row in raw_cats:
        cat_id = to_int(row.get("id"), "Categories.id")
        if cat_id is None:
            print(
                f"Warning: Skipping category row with missing ID: {row}",
                file=sys.stderr,
            )
            continue
        name = str(row.get("name") or "").strip()
        if not name:
            raise ValueError(f"Category {cat_id}: 'name' is required.")
        slug = str(row.get("slug") or "").strip() or slugify(name)
        cat = {
            "id": cat_id,
            "name": name,
            "slug": slug,
            "description": str(row.get("description") or ""),
            "parent_id": int(row["parent_id"]) if row.get("parent_id") else None,
            "subcategories": [],
        }
        categories.append(cat)
        cat_lookup[cat_id] = {"id": cat_id, "name": name, "slug": slug}

    # ── inventory keyed by variant_id ─────────────────────────────────────────
    inv_lookup = {}
    for row in raw_inv:
        vid = to_int(row.get("variant_id"), "Inventory.variant_id")
        if vid is None:
            continue
        raw_qty = row.get("quantity")
        qty = to_int(raw_qty, "Inventory.quantity") or 0

        inv_lookup[vid] = {
            "quantity": qty,
            "location": str(row.get("location") or ""),
            "last_updated": str(row.get("last_updated") or ""),
            "_qty": qty,
        }

    # ── variants grouped by product_id ───────────────────────────────────────
    var_by_product = {}
    for row in raw_vars:
        vid = to_int(row.get("id"), "Variants.id")
        if vid is None:
            print(
                f"Warning: Skipping variant row with missing ID: {row}", file=sys.stderr
            )
            continue
        pid = to_int(row.get("product_id"), "Variants.product_id")
        if pid is None:
            print(
                f"Warning: Skipping variant {vid} with missing product_id",
                file=sys.stderr,
            )
            continue
        sku = str(row.get("sku") or "").strip()
        if not sku:
            raise ValueError(f"Variant {vid}: 'sku' is required.")

        threshold = int(row.get("stock_threshold") or 10)
        inv = inv_lookup.get(
            vid, {"quantity": 0, "location": "", "last_updated": "", "_qty": 0}
        )
        inv["is_low_stock"] = inv["_qty"] < threshold

        variant = {
            "id": vid,
            "sku": sku,
            "price": to_float(row.get("price"), "Variants.price"),
            "unit_type": str(row.get("unit_type") or "piece").strip(),
            "attributes": parse_attributes(row.get("attributes")),
            "stock_threshold": threshold,
            "item_code": str(row.get("item_code") or ""),
            "tax_type": str(row.get("tax_type") or "A").strip(),
            "inv": {
                "quantity": inv["quantity"],
                "location": inv["location"],
                "last_updated": inv["last_updated"],
                "is_low_stock": inv["is_low_stock"],
            },
            "variant_images": [],
        }
        var_by_product.setdefault(pid, []).append(variant)

    # ── products ──────────────────────────────────────────────────────────────
    products = []
    for row in raw_prods:
        pid = to_int(row.get("id"), "Products.id")
        if pid is None:
            print(
                f"Warning: Skipping product row with missing ID: {row}", file=sys.stderr
            )
            continue
        name = str(row.get("name") or "").strip()
        cat_id = to_int(row.get("category_id"), "Products.category_id")

        if not name:
            raise ValueError(f"Product {pid}: 'name' is required.")
        if cat_id not in cat_lookup:
            print(
                f"Warning: Skipping product '{name}' (ID {pid}) references unknown category_id {cat_id}.",
                file=sys.stderr,
            )
            continue

        slug = str(row.get("slug") or "").strip() or slugify(name)
        products.append(
            {
                "id": pid,
                "name": name,
                "slug": slug,
                "description": str(row.get("description") or ""),
                "category": cat_lookup[cat_id],
                "base_price": to_float(row.get("base_price"), "Products.base_price"),
                "is_active": parse_bool(row.get("is_active")),
                "product_images": [],
                "variants": var_by_product.get(pid, []),
            }
        )

    return {
        "categories": categories,
        "products": products,
        "metadata": {
            "timestamp": int(time.time()),
            "total_products": len(products),
            "total_categories": len(categories),
        },
    }


@shared_task(bind=True)
def process_xlsx_import_task(self, file_path):
    """
    Celery task to process XLSX import.
    Updates ImportHistory status, runs conversion, then bulk-creates/updates ORM objects.
    """
    try:
        history = ImportHistory.objects.get(task_id=self.request.id)
    except ImportHistory.DoesNotExist:
        raise ValueError(f"ImportHistory not found for task_id={self.request.id}")

    history.status = "PROCESSING"
    history.save()

    try:
        # FIX 1: Pass `file_path` instead of the undefined `storage_key`
        manifest = convert(file_path)

        from django.db import transaction

        # Counters for reporting
        created = {"categories": 0, "products": 0, "variants": 0, "inventory": 0}
        updated = {"categories": 0, "products": 0, "variants": 0, "inventory": 0}

        with transaction.atomic():
            # ── Categories: lookup by name (unique) ──────────────────────────
            cat_id_map = {}  # sheet id → real DB pk
            for cat_data in manifest["categories"]:
                cat_obj, cat_created = Category.objects.update_or_create(
                    name=cat_data["name"],
                    defaults={
                        "slug": cat_data["slug"],
                        "description": cat_data.get("description", ""),
                    },
                )
                cat_id_map[cat_data["id"]] = cat_obj.pk
                if cat_created:
                    created["categories"] += 1
                else:
                    updated["categories"] += 1

            # Second pass: resolve parent_id using real DB pks
            for cat_data in manifest["categories"]:
                if cat_data.get("parent_id") is not None:
                    real_pk = cat_id_map.get(cat_data["id"])
                    parent_pk = cat_id_map.get(cat_data["parent_id"])
                    if real_pk and parent_pk:
                        Category.objects.filter(pk=real_pk).update(parent_id=parent_pk)

            # ── Products: lookup by slug (unique) ─────────────────────────────
            for prod_data in manifest["products"]:
                real_cat_pk = cat_id_map.get(prod_data["category"]["id"])
                if not real_cat_pk:
                    continue
                prod_obj, prod_created = Product.objects.update_or_create(
                    slug=prod_data["slug"],
                    defaults={
                        "name": prod_data["name"],
                        "description": prod_data.get("description", ""),
                        "category_id": real_cat_pk,
                        "base_price": prod_data["base_price"],
                        "is_active": prod_data["is_active"],
                    },
                )
                if prod_created:
                    created["products"] += 1
                else:
                    updated["products"] += 1

                # ── Variants: lookup by sku (unique) ──────────────────────────
                for var_data in prod_data.get("variants", []):
                    variant_obj, var_created = Variant.objects.update_or_create(
                        sku=var_data["sku"],
                        defaults={
                            "product_id": prod_obj.pk,
                            "price": var_data["price"],
                            "unit_type": var_data.get("unit_type", "piece"),
                            "attributes": var_data.get("attributes", {}),
                            "stock_threshold": var_data.get("stock_threshold", 10),
                            "item_code": var_data.get("item_code", ""),
                            "tax_type": var_data.get("tax_type", "A"),
                        },
                    )
                    if var_created:
                        created["variants"] += 1
                    else:
                        updated["variants"] += 1

                    # ── Inventory: 1-to-1 with variant ────────────────────────
                    inv_data = var_data.get("inv", {})
                    inv_obj, inv_created = Inventory.objects.update_or_create(
                        variant=variant_obj,
                        defaults={
                            "quantity": inv_data.get("quantity", 0),
                            "location": inv_data.get("location", ""),
                        },
                    )
                    if inv_created:
                        created["inventory"] += 1
                    else:
                        updated["inventory"] += 1

        # Attach summary to history
        summary = f"created: {created}, updated: {updated}"
        history.error_msg = summary

        # Bust stale caches
        from products.utils.cache_keys import (
            invalidate_category_cache,
            invalidate_product_cache,
            invalidate_variant_cache,
        )

        invalidate_product_cache()
        invalidate_category_cache()
        invalidate_variant_cache()
        
        for role in ("admin", "staff", "cashier", "customer"):
            cache.delete(f"products:dump:role:{role}")

        history.status = "COMPLETED"
        history.save()
        return "done"

    except Exception as exc:
        history.status = "FAILED"
        history.error_msg = str(exc)
        history.save()
        raise


@shared_task
def keep_neon_alive():
    """Prevent Neon database from suspending."""
    from django.db import connection
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
        logger.info("Neon keep-alive: OK")
        return {"alive": True}
    except Exception as e:
        logger.error(f"Neon keep-alive failed: {e}")
        return {"alive": False, "error": str(e)}


@shared_task
def scheduled_cleanup():
    """Example scheduled task - add your cleanup logic."""
    logger.info("Running scheduled cleanup")
    return {"cleaned": True}